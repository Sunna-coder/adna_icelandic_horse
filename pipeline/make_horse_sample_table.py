"""
Generate ancient horse sample table as Excel file.
Source: background/data/Hores_sites_updated.xlsx (humans repo)
Output: background/horse_ancient_samples.xlsx (horse repo)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC = "/home/sigridse/projects/anda_icelandic_humans/background/data/Hores_sites_updated.xlsx"
OUT = "/home/sigridse/projects/adna_icelandic_horse/background/horse_ancient_samples.xlsx"

# Load source
wb_src = openpyxl.load_workbook(SRC)
ws_src = wb_src.active

# Collect rows
rows = []
for r in range(2, ws_src.max_row + 1):
    site    = ws_src.cell(r, 2).value
    if not site:
        continue
    status  = ws_src.cell(r, 1).value
    n       = ws_src.cell(r, 6).value
    museum  = ws_src.cell(r, 8).value
    period  = ws_src.cell(r, 12).value
    cat     = ws_src.cell(r, 13).value
    county  = ws_src.cell(r, 4).value
    notes   = str(ws_src.cell(r, 10).value or "") + " " + str(ws_src.cell(r, 11).value or "")
    notes   = notes.strip()

    # Flag kuml associations
    kuml_flag = ""
    kuml_keywords = ["kuml", "burial", "grave", "dys ii", "dys xiii", "hrossabein úr tveimur kumlum"]
    if any(k in notes.lower() for k in kuml_keywords):
        kuml_flag = "yes"
    if any(k in str(museum or "").lower() for k in ["kuml", "dys"]):
        kuml_flag = "yes"

    rows.append((site.strip(), period or "", cat or "", n or "", museum or "",
                 status or "", kuml_flag, county or "", notes))

# Summary counts
def parse_n(val):
    try:
        return int(str(val).replace("+", "").strip())
    except:
        return 0

settlement_n     = sum(parse_n(r[3]) for r in rows if r[2] == "settlement" and r[6] == "")
settlement_kuml  = sum(parse_n(r[3]) for r in rows if r[2] == "settlement" and r[6] == "yes")
post_n           = sum(parse_n(r[3]) for r in rows if r[2] == "post-settlement")
undated_n        = sum(parse_n(r[3]) for r in rows if r[2] == "undated")

# Build workbook
wb = openpyxl.Workbook()

# ── Sheet 1: Full sample list ──────────────────────────────────────────────
ws = wb.active
ws.title = "All sites"

hdr_font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
hdr_fill  = PatternFill("solid", fgColor="2F4F8F")
norm_font = Font(name="Calibri", size=10)
wrap      = Alignment(wrap_text=True, vertical="top")
top       = Alignment(vertical="top")

settle_fill   = PatternFill("solid", fgColor="D9EAD3")  # green
post_fill     = PatternFill("solid", fgColor="FCE5CD")  # orange
undated_fill  = PatternFill("solid", fgColor="F3F3F3")  # grey
kuml_fill     = PatternFill("solid", fgColor="FFF2CC")  # yellow

headers = ["Site", "Period (CE)", "Period category", "N specimens",
           "Museum numbers", "Sampling status", "Kuml-associated", "County", "Notes"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws.cell(1, c)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws.freeze_panes = "A2"

for r_idx, (site, period, cat, n, museum, status, kuml, county, notes) in enumerate(rows, start=2):
    vals = [site, period, cat, n, museum, status, kuml, county, notes]
    for c_idx, val in enumerate(vals, start=1):
        cell = ws.cell(r_idx, c_idx, value=val)
        cell.font = norm_font
        cell.alignment = wrap if c_idx in (1, 5, 9) else top
    # Row colour
    if kuml == "yes":
        fill = kuml_fill
    elif cat == "settlement":
        fill = settle_fill
    elif cat == "post-settlement":
        fill = post_fill
    else:
        fill = undated_fill
    for c_idx in range(1, len(headers) + 1):
        ws.cell(r_idx, c_idx).fill = fill

# Column widths
widths = [38, 14, 16, 12, 40, 22, 16, 20, 50]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 2: Summary ────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
ws2.column_dimensions["A"].width = 40
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 16

sum_hdrs = ["Category", "N sites", "N specimens (min)"]
ws2.append(sum_hdrs)
for c in range(1, 4):
    cell = ws2.cell(1, c)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

summary = [
    ("Settlement (<1000 CE), non-kuml",
     sum(1 for r in rows if r[2] == "settlement" and r[6] == ""),
     settlement_n),
    ("Settlement (<1000 CE), kuml-associated",
     sum(1 for r in rows if r[2] == "settlement" and r[6] == "yes"),
     settlement_kuml),
    ("Post-settlement",
     sum(1 for r in rows if r[2] == "post-settlement"),
     post_n),
    ("Undated",
     sum(1 for r in rows if r[2] == "undated"),
     undated_n),
    ("TOTAL",
     len(rows),
     settlement_n + settlement_kuml + post_n + undated_n),
]

fills2 = [settle_fill, kuml_fill, post_fill, undated_fill,
          PatternFill("solid", fgColor="D9D9D9")]

for r_idx, ((label, n_sites, n_spec), fill) in enumerate(zip(summary, fills2), start=2):
    for c_idx, val in enumerate([label, n_sites, n_spec], start=1):
        cell = ws2.cell(r_idx, c_idx, value=val)
        cell.font = Font(name="Calibri", size=11,
                         bold=(label == "TOTAL"))
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left",
                                   vertical="center")

ws2.append([])
ws2.append(["Note: N specimens for post-settlement sites marked '+' are minimums."])
ws2.cell(ws2.max_row, 1).font = Font(name="Calibri", size=10, italic=True)

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Settlement (non-kuml): {settlement_n} specimens across "
      f"{sum(1 for r in rows if r[2]=='settlement' and r[6]=='')} sites")
print(f"Settlement (kuml):     {settlement_kuml} specimens across "
      f"{sum(1 for r in rows if r[2]=='settlement' and r[6]=='yes')} sites")
print(f"Post-settlement:       {post_n}+ specimens across "
      f"{sum(1 for r in rows if r[2]=='post-settlement')} sites")
print(f"Undated:               {undated_n} specimens across "
      f"{sum(1 for r in rows if r[2]=='undated')} sites")
